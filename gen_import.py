import openpyxl
from collections import Counter

# Read source data
wb_src = openpyxl.load_workbook('ekub.xlsx', data_only=True)
ws_src = wb_src['ዕቁብ']

members = []
for row in ws_src.iter_rows(min_row=6, values_only=True):
    spot = row[0]
    name = row[1]
    amount = row[2]
    if spot and name and isinstance(spot, int):
        share = 'full' if amount == 21000 else 'half'
        members.append((name, '', spot, share, ''))

# Create import file
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Members'
ws.append(['Name', 'Phone', 'Spot Number', 'Share', 'Notes'])
for m in members:
    ws.append(list(m))

wb.save('equb_import_ready.xlsx')
print(f'Created equb_import_ready.xlsx with {len(members)} members')

spot_counts = Counter(m[2] for m in members)
shared = [(s, c) for s, c in spot_counts.items() if c > 1]
print(f'Shared spots (half pairs): {shared}')
full = sum(1 for m in members if m[3] == 'full')
half = sum(1 for m in members if m[3] == 'half')
print(f'Full spot members: {full}, Half spot members: {half}')

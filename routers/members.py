from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, Request
from routers.deps import _require_feature
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, selectinload
from sqlalchemy import or_, func, case
from pydantic import BaseModel
from typing import Optional, List
from database import (get_db, Member, MemberSpot, Spot, Settings, Payment, Week, Cycle,
                      NotificationLog, PotTransaction, PotDisbursement, cycle_cfg)
import csv, io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter()


class SpotAssignment(BaseModel):
    spot_id: int
    share: str = "full"           # full | half
    weekly_contribution: float = 21000


import re as _re

_PHONE_RE = _re.compile(r"^\+?\d{7,15}$")

def _validate_phone(v: Optional[str]) -> Optional[str]:
    if v is None:
        return None
    cleaned = _re.sub(r"[\s\-()]", "", v)
    if cleaned and not _PHONE_RE.match(cleaned):
        raise ValueError(
            "Phone must be a valid number (e.g. +251912345678). "
            "Only digits, +, spaces, dashes and parentheses allowed."
        )
    return cleaned or None


class MemberCreate(BaseModel):
    name: str
    phone: Optional[str] = None
    spots: List[SpotAssignment] = []
    notes: Optional[str] = None

    def model_post_init(self, __context):
        self.phone = _validate_phone(self.phone)
        if not self.name or not self.name.strip():
            raise ValueError("Name is required")
        self.name = self.name.strip()


class MemberUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    status: Optional[str] = None
    notes: Optional[str] = None

    def model_post_init(self, __context):
        if self.phone is not None:
            self.phone = _validate_phone(self.phone)
        if self.name is not None:
            self.name = self.name.strip()


class SpotAdd(BaseModel):
    spot_id: int
    share: str = "full"
    weekly_contribution: float = 21000


def _member_dict(m: Member, cycle_id: Optional[int] = None) -> dict:
    # Filter assignments by cycle if provided; otherwise fall back to any active (legacy)
    if cycle_id is not None:
        active_sa = [sa for sa in m.spot_assignments if sa.is_active and sa.cycle_id == cycle_id]
    else:
        active_sa = [sa for sa in m.spot_assignments if sa.is_active]

    assignments = [
        {
            "id": sa.id,
            "spot_id": sa.spot_id,
            "spot_number": sa.spot.number if sa.spot else None,
            "spot_type": sa.spot.spot_type if sa.spot else None,
            "share": sa.share,
            "weekly_contribution": sa.weekly_contribution,
            "cycle_id": sa.cycle_id,
        }
        for sa in active_sa
    ]
    total_weekly = sum(a["weekly_contribution"] for a in assignments)
    spot_numbers = [a["spot_number"] for a in assignments]

    # Find half-spot partners within the same cycle
    partners = []
    for sa in active_sa:
        if sa.share == "half" and sa.spot:
            for other_sa in sa.spot.spot_assignments:
                if (other_sa.member_id != m.id and other_sa.is_active
                        and (cycle_id is None or other_sa.cycle_id == cycle_id)):
                    partners.append(other_sa.member.name)

    return {
        "id": m.id,
        "name": m.name,
        "phone": m.phone,
        "status": m.status,
        "spots": assignments,
        "spot_numbers": spot_numbers,
        "total_weekly_contribution": total_weekly,
        "spot_count": len(assignments),
        "partners": list(set(partners)),
        "notes": m.notes,
        "created_at": m.created_at.isoformat() if m.created_at else None,
        "updated_at": m.updated_at.isoformat() if m.updated_at else None,
        "deleted_at": m.deleted_at.isoformat() if getattr(m, "deleted_at", None) else None,
    }


@router.get("")
def list_members(search: str = "", status: str = "",
                 cycle_id: Optional[int] = None,
                 limit: int = 500, offset: int = 0,
                 db: Session = Depends(get_db)):
    q = db.query(Member)
    if search:
        try:
            # Numeric input → exact spot number match only (no name/phone mixing)
            spot_num = int(search.strip())
            spot_match = db.query(Spot).filter(Spot.number == spot_num).first()
            spot_member_ids = []
            if spot_match:
                spot_member_ids = [
                    ms.member_id for ms in spot_match.spot_assignments
                    if ms.is_active and (cycle_id is None or ms.cycle_id == cycle_id)
                ]
            q = q.filter(Member.id.in_(spot_member_ids)) if spot_member_ids else q.filter(Member.id == -1)
        except ValueError:
            # Text input → name or phone search
            q = q.filter(or_(Member.name.ilike(f"%{search}%"), Member.phone.ilike(f"%{search}%")))
    if status:
        q = q.filter(Member.status == status)
    if cycle_id:
        # Only members who have at least one active assignment in this cycle
        member_ids_in_cycle = [
            r[0] for r in db.query(MemberSpot.member_id).filter(
                MemberSpot.cycle_id == cycle_id,
                MemberSpot.is_active == True,
            ).distinct().all()
        ]
        q = q.filter(Member.id.in_(member_ids_in_cycle))
    members = (q.options(
        selectinload(Member.spot_assignments).options(
            selectinload(MemberSpot.spot).selectinload(
                Spot.spot_assignments
            ).selectinload(MemberSpot.member)
        )
    ).order_by(Member.name).offset(offset).limit(limit).all())
    return [_member_dict(m, cycle_id) for m in members]


@router.get("/stats")
def member_stats(cycle_id: Optional[int] = None, db: Session = Depends(get_db)):
    status_agg = db.query(
        func.count(Member.id).label("total"),
        func.sum(case((Member.status == "active",   1), else_=0)).label("active"),
        func.sum(case((Member.status == "received", 1), else_=0)).label("received"),
        func.sum(case((Member.status == "left",     1), else_=0)).label("left"),
    )
    if cycle_id:
        member_ids_sq = (
            db.query(MemberSpot.member_id)
            .filter(MemberSpot.cycle_id == cycle_id, MemberSpot.is_active == True)
            .distinct()
        )
        row = status_agg.filter(Member.id.in_(member_ids_sq)).one()
        total_spots_assigned = db.query(MemberSpot).filter(
            MemberSpot.cycle_id == cycle_id, MemberSpot.is_active == True
        ).count()
    else:
        row = status_agg.one()
        total_spots_assigned = db.query(MemberSpot).filter(MemberSpot.is_active == True).count()

    return {
        "total": row.total or 0,
        "active": row.active or 0,
        "received": row.received or 0,
        "left": row.left or 0,
        "total_spots_assigned": total_spots_assigned,
    }


@router.get("/available-spots")
def available_spots(db: Session = Depends(get_db)):
    # Scope availability to the active cycle's memberships
    cycle = db.query(Cycle).filter(Cycle.status == "active").first()
    gs = db.query(Settings).first()
    cfg = cycle_cfg(cycle, gs)
    n_total = (cfg.total_member_spots or 0) + (cfg.total_assoc_spots or 0) if cycle else 99999
    spots = (db.query(Spot)
             .filter(Spot.status == "active", Spot.number <= n_total)
             .options(selectinload(Spot.spot_assignments).selectinload(MemberSpot.member))
             .order_by(Spot.number).all())
    result = []
    for s in spots:
        # Only consider assignments belonging to the active cycle (or legacy NULL)
        if cycle:
            active_assignments = [sa for sa in s.spot_assignments
                                  if sa.is_active and sa.cycle_id == cycle.id]
        else:
            active_assignments = [sa for sa in s.spot_assignments if sa.is_active]

        # A full-share occupant locks the spot exclusively — not available to anyone else
        has_full_occupant = any(sa.share == "full" for sa in active_assignments)
        if has_full_occupant:
            continue

        # Only half-share occupants remain
        half_count = len(active_assignments)
        if half_count < 2:
            result.append({
                "id": s.id,
                "number": s.number,
                "type": s.spot_type,
                "occupants": half_count,
                "is_half_available": half_count == 1,
                "current_members": [sa.member.name for sa in active_assignments],
            })
    return result


# ── Apply current settings to active cycle spots ─────────────────────────────

@router.post("/apply-settings-to-cycle")
def apply_settings_to_cycle(request: Request, db: Session = Depends(get_db)):
    """Push global Settings amounts into the active cycle's own settings and all active
    MemberSpot records. Use after changing global spot amounts for the current cycle."""
    _require_feature(request, db, "manage_members")
    settings = db.query(Settings).first()
    if not settings:
        raise HTTPException(status_code=500, detail="Settings not configured")
    cycle = db.query(Cycle).filter(Cycle.status == "active").first()
    if not cycle:
        raise HTTPException(status_code=404, detail="No active cycle")

    # Update the cycle's own settings snapshot from global settings
    cycle.full_spot_amount    = settings.full_spot_amount
    cycle.half_spot_amount    = settings.half_spot_amount
    cycle.association_deduction = settings.association_deduction
    cycle.full_spot_voucher   = getattr(settings, 'full_spot_voucher', 80)
    cycle.half_spot_voucher   = getattr(settings, 'half_spot_voucher', 40)

    spots = db.query(MemberSpot).filter(
        MemberSpot.cycle_id == cycle.id,
        MemberSpot.is_active == True,
    ).all()

    full_amt = settings.full_spot_amount or 21000
    half_amt = settings.half_spot_amount or 10500
    updated = 0
    for ms in spots:
        new_amt = full_amt if ms.share == "full" else half_amt
        if ms.weekly_contribution != new_amt:
            ms.weekly_contribution = new_amt
            updated += 1

    # Also recompute existing pending/missed payments for this cycle
    payments_updated = 0
    from database import Payment, Week
    pending_payments = (
        db.query(Payment)
        .join(Week, Week.id == Payment.week_id)
        .filter(Week.cycle_id == cycle.id, Payment.status.in_(["pending", "missed"]))
        .all()
    )
    member_amounts: dict = {}
    for ms in spots:
        member_amounts.setdefault(ms.member_id, 0)
        member_amounts[ms.member_id] += ms.weekly_contribution
    for p in pending_payments:
        correct = member_amounts.get(p.member_id, 0)
        if correct and p.amount != correct:
            p.amount = correct
            payments_updated += 1

    db.commit()
    return {
        "spots_updated": updated,
        "payments_updated": payments_updated,
        "full_spot_amount": full_amt,
        "half_spot_amount": half_amt,
    }


# ── Export (must be before /{member_id} routes) ──────────────────────────────

@router.get("/export")
def export_members(format: str = Query("csv", pattern="^(csv|xlsx)$"),
                   cycle_id: Optional[int] = None,
                   db: Session = Depends(get_db)):
    """Download member list as CSV or Excel. Pass cycle_id to export only that cycle's members."""
    if cycle_id:
        member_ids = [
            r[0] for r in db.query(MemberSpot.member_id).filter(
                MemberSpot.cycle_id == cycle_id, MemberSpot.is_active == True
            ).distinct().all()
        ]
        members = db.query(Member).filter(Member.id.in_(member_ids)).order_by(Member.id).all()
    else:
        members = db.query(Member).order_by(Member.id).all()

    headers_row = ["Name", "Phone", "Status", "Spot Numbers", "Share Types",
                   "Weekly Contribution (ETB)", "Partners", "Notes"]

    def _row(m):
        if cycle_id:
            assignments = [sa for sa in m.spot_assignments if sa.is_active and sa.cycle_id == cycle_id]
        else:
            assignments = [sa for sa in m.spot_assignments if sa.is_active]
        return [
            m.name,
            m.phone or "",
            m.status,
            ", ".join(str(sa.spot.number) for sa in assignments),
            ", ".join(sa.share for sa in assignments),
            sum(sa.weekly_contribution for sa in assignments),
            ", ".join(sorted(set(
                other.member.name
                for sa in assignments if sa.share == "half"
                for other in sa.spot.spot_assignments
                if other.is_active and other.member_id != m.id
                and (not cycle_id or other.cycle_id == cycle_id)
            ))),
            m.notes or "",
        ]

    if format == "csv":
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(headers_row)
        for m in members:
            w.writerow(_row(m))
        content = buf.getvalue().encode("utf-8-sig")
        return StreamingResponse(
            io.BytesIO(content), media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=members_export.csv"},
        )

    # Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members"
    header_fill = PatternFill("solid", fgColor="1a7a4a")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers_row, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for r, m in enumerate(members, 2):
        for col, val in enumerate(_row(m), 1):
            ws.cell(row=r, column=col, value=val)
    col_widths = [30, 18, 12, 16, 14, 26, 30, 24]
    for i, w_val in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w_val
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=members_export.xlsx"},
    )


@router.get("/import-template")
def import_template(db: Session = Depends(get_db)):
    """Download a blank CSV import template with example row."""
    spots = db.query(Spot).filter(Spot.status == "active").order_by(Spot.number).all()
    taken = {sa.spot.number for s in spots for sa in s.spot_assignments if sa.is_active}
    example_spot = next((s.number for s in spots if s.number not in taken), "")
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Name", "Phone", "Spot Number", "Share", "Notes"])
    w.writerow(["Example Member", "+251912345678", example_spot, "full",
                "Example row — delete before importing"])
    content = buf.getvalue().encode("utf-8-sig")
    return StreamingResponse(
        io.BytesIO(content), media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=import_template.csv"},
    )


# ── Import (must be before /{member_id} routes) ───────────────────────────────

@router.post("/import")
async def import_members(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    _require_feature(request, db, "manage_members")
    """Import members from CSV or Excel. Returns per-row results."""
    filename = (file.filename or "").lower()
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(400, "File too large (max 5 MB)")

    rows = []
    try:
        if filename.endswith(".csv"):
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            rows = [r for r in reader if any(v and str(v).strip() for v in r.values())]
        elif filename.endswith((".xlsx", ".xls")):
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            iter_rows = list(ws.iter_rows(values_only=True))
            if not iter_rows:
                raise HTTPException(400, "Empty file")
            headers = [str(h).strip() if h else "" for h in iter_rows[0]]
            for row in iter_rows[1:]:
                if any(cell is not None for cell in row):
                    rows.append(dict(zip(headers,
                        [str(c).strip() if c is not None else "" for c in row])))
        else:
            raise HTTPException(400, "Only .csv or .xlsx files are supported")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Could not parse file: {e}")

    def _get(row, *keys):
        for k in keys:
            for rk in row:
                if rk and rk.strip().lower() == k.lower():
                    v = str(row[rk]).strip() if row[rk] else ""
                    if v:
                        return v
        return ""

    # Get active cycle for spot assignments
    active_cycle = db.query(Cycle).filter(Cycle.status == "active").first()
    if active_cycle and active_cycle.draw_phase == "active":
        raise HTTPException(status_code=400,
            detail="Draws have already started — importing members mid-cycle is not allowed.")
    active_cycle_id = active_cycle.id if active_cycle else None
    gs  = db.query(Settings).first()
    cfg = cycle_cfg(active_cycle, gs)

    results = []
    created = 0
    skipped = 0

    for i, row in enumerate(rows, 1):
        name     = _get(row, "Name", "Full Name", "member name", "member")
        phone    = _get(row, "Phone", "Phone Number", "mobile", "tel")
        spot_str = _get(row, "Spot Number", "Spot", "spot_number", "spot no")
        share    = (_get(row, "Share", "Share Type", "share_type") or "full").lower().strip()
        notes    = _get(row, "Notes", "note", "remark")

        if share not in ("full", "half"):
            share = "full"

        if not name:
            results.append({"row": i, "status": "skipped", "reason": "Empty name"})
            skipped += 1
            continue

        m = Member(name=name, phone=phone or None, notes=notes or None)
        db.add(m)
        db.flush()

        spot_msg = None
        if spot_str:
            try:
                spot_num = int(float(spot_str))
                spot = db.query(Spot).filter(Spot.number == spot_num).first()
                if not spot:
                    spot_msg = f"Spot #{spot_num} not found — created without spot"
                else:
                    # Count only within the active cycle, not old cycles
                    active_count = sum(
                        1 for sa in spot.spot_assignments
                        if sa.is_active and sa.cycle_id == active_cycle_id
                    )
                    if active_count >= 2:
                        spot_msg = f"Spot #{spot_num} is full — created without spot"
                    else:
                        if active_count == 1 and share == "full":
                            share = "half"
                            spot_msg = f"Spot #{spot_num} has 1 occupant — assigned as half"
                        else:
                            spot_msg = f"Spot #{spot_num} ({share})"
                        contribution = (
                            cfg.full_spot_amount if share == "full" else cfg.half_spot_amount
                        )
                        db.add(MemberSpot(member_id=m.id, spot_id=spot.id,
                                          share=share, weekly_contribution=contribution,
                                          cycle_id=active_cycle_id))
            except (ValueError, TypeError):
                spot_msg = f"Invalid spot number '{spot_str}' — created without spot"

        results.append({
            "row": i, "status": "created",
            "name": name, "phone": phone or None,
            "spot": spot_msg,
            "warning": ("⚠ " + spot_msg) if spot_msg and "without" in spot_msg else None,
        })
        created += 1

    db.commit()
    return {"total": len(rows), "created": created, "skipped": skipped, "results": results}


# ── CRUD (parameterised routes — must be AFTER all fixed-path routes) ─────────

@router.delete("/permanent/all")
def delete_all_members(request: Request, db: Session = Depends(get_db)):
    """Wipe all members. Admin only.
    Members with no payment records are hard-deleted.
    Members that have payment records are soft-deleted (marked as left).
    """
    _require_feature(request, db, "manage_members")

    # Remove all spot assignments first
    db.query(MemberSpot).delete(synchronize_session=False)
    db.flush()

    all_members = db.query(Member).all()
    hard_deleted = 0
    soft_deleted = 0
    for m in all_members:
        pay_count = db.query(Payment).filter(Payment.member_id == m.id).count()
        if pay_count == 0:
            db.delete(m)
            hard_deleted += 1
        else:
            m.status = "left"
            soft_deleted += 1

    db.commit()
    return {"ok": True, "hard_deleted": hard_deleted, "soft_deleted": soft_deleted,
            "message": f"{hard_deleted} permanently deleted, {soft_deleted} marked as left (had payment records)"}


@router.delete("/permanent/{member_id}")
def delete_member_permanent(member_id: int, request: Request, db: Session = Depends(get_db)):
    """Hard-delete a member and all their records. Admin only."""
    _require_feature(request, db, "manage_members")
    m = db.query(Member).filter(Member.id == member_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Member not found")

    # Block deletion if member is a guarantor on any cheque — deleting would destroy the audit trail
    from sqlalchemy import or_
    is_guarantor = db.query(PotDisbursement).filter(
        or_(
            PotDisbursement.guarantor_1_id == member_id,
            PotDisbursement.guarantor_2_id == member_id,
            PotDisbursement.guarantor_3_id == member_id,
        )
    ).first()
    if is_guarantor:
        raise HTTPException(
            status_code=409,
            detail=(
                f"{m.name} is listed as a guarantor on one or more cheque records. "
                "Re-record those disbursements with a different guarantor before deleting this member."
            ),
        )

    # 1. Notification logs
    db.query(NotificationLog).filter(
        NotificationLog.member_id == member_id
    ).delete(synchronize_session=False)

    # 2. Pot transactions: null out nullable refs; delete rows where buyer (NOT NULL)
    db.query(PotTransaction).filter(
        PotTransaction.buyer_id == member_id
    ).delete(synchronize_session=False)
    db.query(PotTransaction).filter(
        PotTransaction.seller_id == member_id
    ).update({"seller_id": None}, synchronize_session=False)
    db.query(PotTransaction).filter(
        PotTransaction.original_winner_id == member_id
    ).update({"original_winner_id": None}, synchronize_session=False)

    # 3. Payments
    db.query(Payment).filter(
        Payment.member_id == member_id
    ).delete(synchronize_session=False)

    # 4. Spot assignments
    db.query(MemberSpot).filter(
        MemberSpot.member_id == member_id
    ).delete(synchronize_session=False)

    db.flush()
    db.delete(m)
    db.commit()
    return {"ok": True}


@router.post("")
def create_member(data: MemberCreate, request: Request, db: Session = Depends(get_db)):
    _require_feature(request, db, "manage_members")
    cycle = db.query(Cycle).filter(Cycle.status == "active").first()
    if cycle and cycle.draw_phase == "active":
        raise HTTPException(status_code=400,
            detail="Draws have already started — new members cannot be added mid-cycle.")
    gs    = db.query(Settings).first()
    cfg   = cycle_cfg(cycle, gs)
    try:
        m = Member(name=data.name, phone=data.phone, notes=data.notes)
        db.add(m)
        db.flush()

        for assignment in data.spots:
            spot = db.query(Spot).filter(Spot.id == assignment.spot_id).first()
            if not spot:
                raise HTTPException(status_code=404, detail=f"Spot {assignment.spot_id} not found")
            cycle_id_val = cycle.id if cycle else None
            active_assignments = [sa for sa in spot.spot_assignments
                                  if sa.is_active and sa.cycle_id == cycle_id_val]
            has_full = any(sa.share == "full" for sa in active_assignments)
            half_count = sum(1 for sa in active_assignments if sa.share == "half")

            if has_full:
                raise HTTPException(status_code=400,
                    detail=f"Spot #{spot.number} is fully occupied by a full-spot member — it cannot be shared.")
            if assignment.share == "full" and half_count > 0:
                raise HTTPException(status_code=400,
                    detail=f"Spot #{spot.number} already has a half-spot member — assign a different spot for a full spot.")
            if assignment.share == "half" and half_count >= 2:
                raise HTTPException(status_code=400,
                    detail=f"Spot #{spot.number} is full (2 half-spot members already).")
            # Enforce contribution amount from this cycle's settings
            contribution = (
                cfg.full_spot_amount if assignment.share == "full" else cfg.half_spot_amount
            )
            db.add(MemberSpot(
                member_id=m.id, spot_id=assignment.spot_id,
                share=assignment.share, weekly_contribution=contribution,
                cycle_id=cycle.id if cycle else None,
            ))

        db.commit()
        db.refresh(m)
        return _member_dict(m)
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create member: {e}")


@router.get("/{member_id}")
def get_member(member_id: int, db: Session = Depends(get_db)):
    m = db.query(Member).filter(Member.id == member_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Member not found")
    return _member_dict(m)


@router.put("/{member_id}")
def update_member(member_id: int, data: MemberUpdate, request: Request, db: Session = Depends(get_db)):
    _require_feature(request, db, "manage_members")
    m = db.query(Member).filter(Member.id == member_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Member not found")
    for field, value in data.model_dump(exclude_none=True).items():
        setattr(m, field, value)
    db.commit()
    db.refresh(m)
    return _member_dict(m)


@router.post("/{member_id}/spots")
def add_spot(member_id: int, data: SpotAdd, request: Request, db: Session = Depends(get_db)):
    _require_feature(request, db, "manage_members")
    m = db.query(Member).filter(Member.id == member_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Member not found")
    spot = db.query(Spot).filter(Spot.id == data.spot_id).first()
    if not spot:
        raise HTTPException(status_code=404, detail="Spot not found")

    cycle = db.query(Cycle).filter(Cycle.status == "active").first()
    if cycle and cycle.draw_phase == "active":
        raise HTTPException(status_code=400,
            detail="Draws have already started — spot assignments cannot be changed mid-cycle.")
    gs    = db.query(Settings).first()
    cfg   = cycle_cfg(cycle, gs)
    contribution = (
        cfg.full_spot_amount if data.share == "full" else cfg.half_spot_amount
    )

    existing = db.query(MemberSpot).filter_by(member_id=member_id, spot_id=data.spot_id,
                                               cycle_id=cycle.id if cycle else None).first()
    if existing:
        if existing.is_active:
            raise HTTPException(status_code=400, detail="Member already assigned to this spot")
        existing.is_active = True
        existing.share = data.share
        existing.weekly_contribution = contribution
    else:
        cycle_id_val = cycle.id if cycle else None
        active_assignments = [sa for sa in spot.spot_assignments
                              if sa.is_active and sa.member_id != member_id
                              and sa.cycle_id == cycle_id_val]
        has_full = any(sa.share == "full" for sa in active_assignments)
        half_count = sum(1 for sa in active_assignments if sa.share == "half")

        if has_full:
            raise HTTPException(status_code=400,
                detail=f"Spot #{spot.number} is fully occupied by a full-spot member — it cannot be shared.")
        if data.share == "full" and half_count > 0:
            raise HTTPException(status_code=400,
                detail=f"Spot #{spot.number} already has a half-spot member — choose a different spot for a full spot.")
        if data.share == "half" and half_count >= 2:
            raise HTTPException(status_code=400,
                detail=f"Spot #{spot.number} is full (2 half-spot members already).")
        db.add(MemberSpot(
            member_id=member_id, spot_id=data.spot_id,
            share=data.share, weekly_contribution=contribution,
            cycle_id=cycle.id if cycle else None,   # ← scoped to active cycle
        ))

    db.commit()
    db.refresh(m)
    return _member_dict(m)


@router.delete("/{member_id}/spots/{spot_id}")
def remove_spot(member_id: int, spot_id: int, request: Request, db: Session = Depends(get_db)):
    _require_feature(request, db, "manage_members")
    active_cycle = db.query(Cycle).filter(Cycle.status == "active").first()
    cycle_id = active_cycle.id if active_cycle else None
    # Prefer the active-cycle-scoped assignment; fall back to any matching row (legacy data)
    sa = (
        db.query(MemberSpot)
        .filter(MemberSpot.member_id == member_id,
                MemberSpot.spot_id == spot_id,
                MemberSpot.cycle_id == cycle_id)
        .first()
    )
    if not sa:
        sa = (db.query(MemberSpot)
              .filter(MemberSpot.member_id == member_id, MemberSpot.spot_id == spot_id)
              .first())
    if not sa:
        raise HTTPException(status_code=404, detail="Spot assignment not found")
    sa.is_active = False
    db.commit()
    return {"ok": True}


@router.delete("/{member_id}")
def mark_left(member_id: int, request: Request, db: Session = Depends(get_db)):
    _require_feature(request, db, "manage_members")
    m = db.query(Member).filter(Member.id == member_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Member not found")
    from datetime import datetime, timezone
    m.status = "left"
    m.deleted_at = datetime.now(timezone.utc).replace(tzinfo=None)
    for sa in m.spot_assignments:
        sa.is_active = False
    db.commit()
    return {"ok": True}


class MemberExitIn(BaseModel):
    exit_week_id: int
    reason: str = "left"          # left | stopped_paying


@router.post("/{member_id}/exit")
def member_exit(member_id: int, data: MemberExitIn, request: Request,
                db: Session = Depends(get_db)):
    """
    Record a member exit at a specific week.
    - Marks member status = left
    - Deactivates their MemberSpot, records exit week + reason
    - Marks all pending/late payments AFTER the exit week as missed
    - Returns a financial summary
    """
    _require_feature(request, db, "manage_members")
    m = db.query(Member).filter(Member.id == member_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Member not found")

    exit_week = db.query(Week).filter(Week.id == data.exit_week_id).first()
    if not exit_week:
        raise HTTPException(status_code=404, detail="Exit week not found")

    cycle_id = exit_week.cycle_id

    # Mark member left
    m.status = "left"

    # Deactivate spot assignments for this cycle, record exit info
    for sa in m.spot_assignments:
        if sa.cycle_id == cycle_id:
            sa.is_active = False
            sa.exited_at_week_id = data.exit_week_id
            sa.exit_reason = data.reason

    # All pending/late payments after the exit week → missed
    future_week_ids = [
        r[0] for r in db.query(Week.id).filter(
            Week.cycle_id == cycle_id,
            Week.week_number > exit_week.week_number
        ).all()
    ]
    if future_week_ids:
        db.query(Payment).filter(
            Payment.member_id == member_id,
            Payment.week_id.in_(future_week_ids),
            Payment.status.in_(["pending", "late"])
        ).update({"status": "missed"}, synchronize_session=False)

    db.commit()

    # ── Financial summary ─────────────────────────────────────────────────────
    all_pmts = db.query(Payment).filter(
        Payment.member_id == member_id,
        Payment.week_id.in_(
            [r[0] for r in db.query(Week.id).filter(Week.cycle_id == cycle_id).all()]
        )
    ).all()

    paid_weeks   = [p for p in all_pmts if p.status == "paid"]
    missed_weeks = [p for p in all_pmts if p.status == "missed"]
    total_paid   = sum(p.amount for p in paid_weeks)
    total_missed = sum(p.amount for p in missed_weeks)

    # Did this member ever win the pot in this cycle?
    has_won = db.query(PotDisbursement).join(Week).filter(
        Week.cycle_id == cycle_id,
        PotDisbursement.member_id == member_id
    ).first() is not None

    # Also check via spot draw
    if not has_won:
        spots = [sa.spot_id for sa in m.spot_assignments if sa.cycle_id == cycle_id]
        has_won = db.query(Week).filter(
            Week.cycle_id == cycle_id,
            Week.winner_spot_id.in_(spots),
            Week.status.in_(["drawn", "sold"])
        ).first() is not None

    return {
        "ok": True,
        "member_id": member_id,
        "member_name": m.name,
        "exit_week_number": exit_week.week_number,
        "reason": data.reason,
        "weeks_paid": len(paid_weeks),
        "total_paid": total_paid,
        "weeks_missed": len(missed_weeks),
        "total_missed": total_missed,
        "has_won": has_won,
        "summary": (
            f"Won the pot — still owes {len(missed_weeks)} week(s) = {total_missed:,.0f} ETB"
            if has_won and missed_weeks else
            "Won the pot — fully settled" if has_won else
            f"Left without winning — {len(paid_weeks)} week(s) paid ({total_paid:,.0f} ETB)"
        ),
    }
